import os, re, secrets
from datetime import datetime, timedelta
from functools import wraps

import bcrypt
from dotenv import load_dotenv
from flask import (Flask, request, jsonify, session, redirect,
                   url_for, render_template, send_file, abort, g)
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_mail import Mail, Message
from flask_jwt_extended import (JWTManager, create_access_token,
                                 jwt_required, get_jwt_identity)
from flask_wtf.csrf import CSRFProtect
from werkzeug.middleware.proxy_fix import ProxyFix
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from werkzeug.utils import secure_filename
import tempfile

# ── bootstrap ──────────────────────────────────────────────────────────────
load_dotenv()

from database import get_db, init_db, COLLEGES, ROLES, log_action, ph, USE_PG
from image_processor import (detect_faces, apply_edits, face_detected, save_image,
                              archive_old_image, TARGET_W, TARGET_H, validate_single_person)

app = Flask(__name__)
# Enable ProxyFix behind a reverse proxy (e.g. Nginx)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1)

# Initialize CSRF Protection
csrf = CSRFProtect(app)

# Load configuration securely
secret_key_env = os.getenv("SECRET_KEY")
if not secret_key_env:
    app.logger.warning("[SECURITY WARNING] SECRET_KEY not set. Using temporary random key.")
    secret_key_env = secrets.token_hex(32)
app.secret_key = secret_key_env

jwt_secret_env = os.getenv("JWT_SECRET_KEY")
if not jwt_secret_env:
    app.logger.warning("[SECURITY WARNING] JWT_SECRET_KEY not set. Using temporary random key.")
    jwt_secret_env = secrets.token_hex(32)
app.config["JWT_SECRET_KEY"]           = jwt_secret_env

app.config["JWT_ACCESS_TOKEN_EXPIRES"] = timedelta(hours=8)
app.config["MAX_CONTENT_LENGTH"]       = int(os.getenv("MAX_CONTENT_LENGTH", 5*1024*1024))
app.config["MAIL_SERVER"]              = os.getenv("MAIL_SERVER",  "smtp.gmail.com")
app.config["MAIL_PORT"]                = int(os.getenv("MAIL_PORT", 587))
app.config["MAIL_USE_TLS"]             = os.getenv("MAIL_USE_TLS","true").lower()=="true"

# SMTP Credentials from Environment Variables (V-001)
app.config["MAIL_USERNAME"]            = os.getenv("MAIL_USERNAME")
app.config["MAIL_PASSWORD"]            = os.getenv("MAIL_PASSWORD")
app.config["MAIL_DEFAULT_SENDER"]      = os.getenv("MAIL_DEFAULT_SENDER")

UPLOAD_FOLDER        = os.path.join(app.root_path, "static", "uploads")
STATIC_ROOT          = os.path.join(app.root_path, "static")
UNIVERSITY_DOMAIN    = os.getenv("UNIVERSITY_EMAIL_DOMAIN", "bua.edu.eg")
CURRENT_YEAR         = datetime.now().year
YEAR_RANGE           = list(range(CURRENT_YEAR, CURRENT_YEAR - 10, -1))

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@app.after_request
def add_security_headers(response):
    response.headers["X-Frame-Options"] = "SAMEORIGIN"  # Clickjacking mitigation (V-008)
    response.headers["X-Content-Type-Options"] = "nosniff"  # MIME sniffing prevention (V-024)
    # Content Security Policy (V-022)
    response.headers["Content-Security-Policy"] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "img-src 'self' data: https://res.cloudinary.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "frame-ancestors 'none';"
    )
    # Strict Transport Security (HSTS) (V-023)
    response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    return response

mail    = Mail(app)
jwt     = JWTManager(app)
limiter = Limiter(
    key_func=get_remote_address, app=app,
    default_limits=["300 per day", "80 per hour"],
    storage_uri="memory://",
)

ARABIC_MAP = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")

# ── helpers ────────────────────────────────────────────────────────────────

def to_eng(s): return s.translate(ARABIC_MAP)

def validate_university_email(email: str) -> bool:
    email = email.strip().lower()
    return re.fullmatch(r"[^@\s]+@" + re.escape(UNIVERSITY_DOMAIN), email) is not None

def validate_full_name(name: str) -> tuple[bool, str]:
    if len(name.strip().split()) < 4:
        return False, "الاسم يجب أن يكون رباعياً على الأقل"
    return True, ""

def validate_student_id(year: str, code: str) -> tuple[bool, str]:
    if not re.fullmatch(r"\d{4}", year):   return False, "السنة يجب أن تكون 4 أرقام"
    if not (CURRENT_YEAR - 10 <= int(year) <= CURRENT_YEAR):
        return False, f"سنة القيد غير صالحة. يجب أن تكون بين {CURRENT_YEAR - 10} و {CURRENT_YEAR}"
    if not re.fullmatch(r"\d{6}|\d{8}", code): return False, "الكود يجب أن يكون 6 أو 8 أرقام"
    return True, ""

def extract_student_id_from_email(email: str) -> str:
    """Extract student ID from email like AbdulRahman.2023006972@bua.edu.eg"""
    email_prefix = email.split("@")[0]
    # Match pattern: anything.10-12digits
    match = re.search(r'\.(\d{10,12})$', email_prefix)
    if match:
        return match.group(1)
    return None

def hash_pw(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()

def check_pw(pw: str, hashed: str) -> bool:
    return bcrypt.checkpw(pw.encode(), hashed.encode())

def send_email(to: str, subject: str, html: str):
    if not app.config.get("MAIL_USERNAME") or not app.config.get("MAIL_PASSWORD"):
        app.logger.warning("📧 MAIL not configured – skipping email to %s | Subject: %s", to, subject)
        app.logger.warning("📧 To fix: set MAIL_USERNAME + MAIL_PASSWORD in .env (use Gmail App Password)")
        return
    try:
        msg = Message(subject, recipients=[to], html=html)
        mail.send(msg)
        app.logger.info("📧 Email sent to %s", to)
    except Exception as e:
        app.logger.error("📧 Email error to %s: %s", to, e)

def _row_to_dict(row) -> dict:
    if row is None: return {}
    if USE_PG: return dict(row)
    return dict(zip(row.keys(), row))

# ── auth decorators ────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def inner(*a, **kw):
        if not session.get("user_id"):
            return redirect(url_for("auth_login"))
        return f(*a, **kw)
    return inner

def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def inner(*a, **kw):
            if not session.get("user_id"):
                return redirect(url_for("auth_login"))
            if session.get("role") not in roles:
                abort(403)
            return f(*a, **kw)
        return inner
    return decorator

def _current_user():
    uid = session.get("user_id")
    if not uid: return None
    db  = get_db(); cur = db.cursor()
    cur.execute(f"SELECT * FROM users WHERE id={ph()}", (uid,))
    row = _row_to_dict(cur.fetchone()); db.close()
    return row

# ── email templates ────────────────────────────────────────────────────────

def _email_verify_html(name, link):
    return f"""
<div dir="rtl" style="font-family:Cairo,Arial;max-width:520px;margin:auto">
  <div style="background:#0d1f3c;padding:28px;border-radius:14px 14px 0 0;text-align:center">
    <h2 style="color:#e8b84b;margin:0">تأكيد البريد الإلكتروني</h2>
  </div>
  <div style="background:#f0f4f9;padding:28px;border-radius:0 0 14px 14px">
    <p>أهلاً <strong>{name}</strong>،</p>
    <p>انقر على الزر أدناه لتفعيل حسابك:</p>
    <a href="{link}" style="display:inline-block;background:#0d1f3c;color:#e8b84b;padding:12px 28px;border-radius:8px;text-decoration:none;font-weight:700;margin:16px 0">تفعيل الحساب</a>
    <p style="color:#888;font-size:.85rem">الرابط صالح لمدة 24 ساعة</p>
  </div>
</div>"""

def _email_reset_html(name, link):
    return f"""
<div dir="rtl" style="font-family:Cairo,Arial;max-width:520px;margin:auto">
  <div style="background:#c53030;padding:28px;border-radius:14px 14px 0 0;text-align:center">
    <h2 style="color:#fff;margin:0">إعادة تعيين كلمة المرور</h2>
  </div>
  <div style="background:#f0f4f9;padding:28px;border-radius:0 0 14px 14px">
    <p>أهلاً <strong>{name}</strong>،</p>
    <p>انقر على الزر أدناه لإعادة تعيين كلمة مرورك:</p>
    <a href="{link}" style="display:inline-block;background:#c53030;color:#fff;padding:12px 28px;border-radius:8px;text-decoration:none;font-weight:700;margin:16px 0">إعادة التعيين</a>
    <p style="color:#888;font-size:.85rem">الرابط صالح لمدة ساعة واحدة فقط. إذا لم تطلب ذلك تجاهل هذا البريد.</p>
  </div>
</div>"""

def _email_registered_html(name, student_id, card_link):
    return f"""
<div dir="rtl" style="font-family:Cairo,Arial;max-width:520px;margin:auto">
  <div style="background:#0d1f3c;padding:28px;border-radius:14px 14px 0 0;text-align:center">
    <h2 style="color:#e8b84b;margin:0">🎓 تم تسجيلك بنجاح</h2>
  </div>
  <div style="background:#f0f4f9;padding:28px;border-radius:0 0 14px 14px">
    <p>أهلاً <strong>{name}</strong>،</p>
    <p>تم تسجيلك في النظام بنجاح. رقمك الجامعي هو:</p>
    <div style="background:#0d1f3c;color:#e8b84b;font-family:monospace;font-size:1.4rem;font-weight:700;padding:14px;border-radius:8px;text-align:center;letter-spacing:3px;margin:16px 0">{student_id}</div>
    <a href="{card_link}" style="display:inline-block;background:#1a3a6b;color:#fff;padding:12px 28px;border-radius:8px;text-decoration:none;font-weight:700">عرض بطاقة الهوية</a>
  </div>
</div>"""

# ══════════════════════════════════════════════════════════════════════════
# AUTH ROUTES
# ══════════════════════════════════════════════════════════════════════════

@app.route("/auth/login", methods=["GET","POST"])
@limiter.limit(lambda: os.getenv("LIMIT_AUTH_LOGIN", "30 per hour"))
def auth_login():
    if session.get("user_id"):
        return redirect(url_for("dashboard"))
    error = None
    if request.method == "POST":
        email = request.form.get("email","").strip().lower()
        pw    = request.form.get("password","")
        db    = get_db(); cur = db.cursor()
        cur.execute(f"SELECT * FROM users WHERE email={ph()}", (email,))
        u = _row_to_dict(cur.fetchone()); db.close()
        if not u:
            error = "البريد الإلكتروني غير مسجل"
        elif not u.get("is_active"):
            error = "الحساب موقوف. تواصل مع مدير النظام"
        elif not u.get("email_verified"):
            error = "يرجى تفعيل بريدك الإلكتروني أولاً"
        elif not check_pw(pw, u["password_hash"]):
            error = "كلمة المرور غير صحيحة"
        else:
            session.clear()
            session["user_id"]    = u["id"]
            session["user_name"]  = u["full_name"]
            session["role"]       = u["role"]
            session["email"]      = u["email"]
            session["college"]    = u.get("college") or ""
            session["student_id"] = u.get("student_id") or ""
            log_action(u["id"], "LOGIN", ip=request.remote_addr)
            # Students → check if registered, go to card or self-register
            if u["role"] == "student":
                sid = u.get("student_id") or ""
                if sid:
                    db2  = get_db(); cur2 = db2.cursor()
                    cur2.execute(f"SELECT student_id FROM students WHERE student_id={ph()}", (sid,))
                    exists = cur2.fetchone(); db2.close()
                    if exists:
                        return redirect(url_for("student_card", student_id=sid))
                # Not registered yet → self-registration page
                return redirect(url_for("student_self_register"))
            return redirect(url_for("dashboard"))
    return render_template("auth_login.html", error=error,
                           domain=UNIVERSITY_DOMAIN)


@app.route("/student/login", methods=["GET","POST"])
@limiter.limit(lambda: os.getenv("LIMIT_STUDENT_LOGIN", "30 per hour"))
def student_login():
    """Dedicated login page for students."""
    if session.get("user_id"):
        return redirect(url_for("dashboard"))
    error = None
    if request.method == "POST":
        email = request.form.get("email","").strip().lower()
        pw    = request.form.get("password","")
        db    = get_db(); cur = db.cursor()
        cur.execute(f"SELECT * FROM users WHERE email={ph()}", (email,))
        u = _row_to_dict(cur.fetchone()); db.close()
        if not u:
            error = "البريد الإلكتروني غير مسجل"
        elif u.get("role") != "student":
            error = "هذه الصفحة مخصصة للطلاب فقط"
        elif not u.get("is_active"):
            error = "الحساب موقوف. تواصل مع مدير النظام"
        elif not u.get("email_verified"):
            error = "يرجى تفعيل بريدك الإلكتروني أولاً"
        elif not check_pw(pw, u["password_hash"]):
            error = "كلمة المرور غير صحيحة"
        else:
            session.clear()
            session["user_id"]    = u["id"]
            session["user_name"]  = u["full_name"]
            session["role"]       = u["role"]
            session["email"]      = u["email"]
            session["college"]    = u.get("college") or ""
            session["student_id"] = u.get("student_id") or ""
            log_action(u["id"], "LOGIN", ip=request.remote_addr)
            # Students → check if registered, go to card or self-register
            sid = u.get("student_id") or ""
            if sid:
                db2  = get_db(); cur2 = db2.cursor()
                cur2.execute(f"SELECT student_id FROM students WHERE student_id={ph()}", (sid,))
                exists = cur2.fetchone(); db2.close()
                if exists:
                    return redirect(url_for("student_card", student_id=sid))
            # Not registered yet → self-registration page
            return redirect(url_for("student_self_register"))
    return render_template("student_login.html", error=error,
                           domain=UNIVERSITY_DOMAIN)


@app.route("/auth/register", methods=["GET","POST"])
@limiter.limit(lambda: os.getenv("LIMIT_AUTH_REGISTER", "30 per hour"))
def auth_register():
    """Public self-registration — always assigns 'student' role.
    Admins/staff accounts are created by superadmin only."""
    error = None
    if request.method == "POST":
        email     = request.form.get("email","").strip().lower()
        pw        = request.form.get("password","")
        pw2       = request.form.get("password2","")
        full_name = request.form.get("full_name","").strip()

        if not validate_university_email(email):
            error = f"يجب استخدام إيمبيل الجامعة (@{UNIVERSITY_DOMAIN})"
        elif len(pw) < 8:
            error = "كلمة المرور يجب أن تكون 8 أحرف على الأقل"
        elif pw != pw2:
            error = "كلمتا المرور غير متطابقتين"
        elif len(full_name.split()) < 2:
            error = "يرجى إدخال الاسم الكامل"
        else:
            # Extract student_id from email prefix (e.g. AbdulRahman.2023006972@bua.edu.eg)
            sid = extract_student_id_from_email(email)

            db  = get_db(); cur = db.cursor()
            cur.execute(f"SELECT id FROM users WHERE email={ph()}", (email,))
            if cur.fetchone():
                db.close()
                error = "هذا البريد مسجل مسبقاً"
            else:
                token  = secrets.token_urlsafe(32)
                hashed = hash_pw(pw)
                # Always student role — admin creates staff/admin accounts separately
                cur.execute(
                    f"INSERT INTO users (email,password_hash,full_name,role,student_id,verify_token) VALUES ({','.join([ph()]*6)})",
                    (email, hashed, full_name, "student", sid, token)
                )
                db.commit(); db.close()
                link = url_for("auth_verify", token=token, _external=True)
                send_email(email, "تفعيل حساب نظام التسجيل", _email_verify_html(full_name, link))
                return render_template("auth_register.html",
                    success="تم إنشاء حسابك! تحقق من بريدك لتفعيل الحساب.",
                    domain=UNIVERSITY_DOMAIN)
    return render_template("auth_register.html", error=error, domain=UNIVERSITY_DOMAIN)


@app.route("/auth/verify/<token>")
def auth_verify(token):
    db = get_db(); cur = db.cursor()
    cur.execute(f"SELECT * FROM users WHERE verify_token={ph()}", (token,))
    u = _row_to_dict(cur.fetchone())
    if not u:
        db.close()
        return render_template("auth_message.html",
            title="رابط غير صالح", msg="رابط التفعيل منتهي أو غير صحيح.", type="error")
    if USE_PG:
        cur.execute("UPDATE users SET email_verified=TRUE, verify_token=NULL WHERE id=%s", (u["id"],))
    else:
        cur.execute("UPDATE users SET email_verified=1, verify_token=NULL WHERE id=?", (u["id"],))
    db.commit(); db.close()
    log_action(u["id"], "EMAIL_VERIFIED", ip=request.remote_addr)
    return render_template("auth_message.html",
        title="تم التفعيل ✓", msg="تم تفعيل حسابك بنجاح. يمكنك الآن تسجيل الدخول.", type="success")


@app.route("/auth/forgot", methods=["GET","POST"])
@limiter.limit(lambda: os.getenv("LIMIT_AUTH_FORGOT", "10 per hour"))
def auth_forgot():
    msg = None
    if request.method == "POST":
        email = request.form.get("email","").strip().lower()
        db    = get_db(); cur = db.cursor()
        cur.execute(f"SELECT * FROM users WHERE email={ph()}", (email,))
        u = _row_to_dict(cur.fetchone())
        if u:
            token   = secrets.token_urlsafe(32)
            expires = (datetime.utcnow() + timedelta(hours=1)).isoformat()
            cur.execute(
                f"UPDATE users SET reset_token={ph()}, reset_expires={ph()} WHERE id={ph()}",
                (token, expires, u["id"])
            )
            db.commit()
            link = url_for("auth_reset", token=token, _external=True)
            send_email(email, "إعادة تعيين كلمة المرور", _email_reset_html(u["full_name"], link))
        db.close()
        msg = "إذا كان البريد مسجلاً ستصلك رسالة خلال دقائق"
    return render_template("auth_forgot.html", msg=msg)


@app.route("/auth/reset/<token>", methods=["GET","POST"])
def auth_reset(token):
    db = get_db(); cur = db.cursor()
    cur.execute(f"SELECT * FROM users WHERE reset_token={ph()}", (token,))
    u = _row_to_dict(cur.fetchone())
    if not u or (u.get("reset_expires") and
                 datetime.fromisoformat(str(u["reset_expires"])) < datetime.utcnow()):
        db.close()
        return render_template("auth_message.html",
            title="رابط منتهي", msg="رابط إعادة التعيين منتهي الصلاحية.", type="error")
    error = None
    if request.method == "POST":
        pw  = request.form.get("password","")
        pw2 = request.form.get("password2","")
        if len(pw) < 8: error = "كلمة المرور يجب أن تكون 8 أحرف على الأقل"
        elif pw != pw2:  error = "كلمتا المرور غير متطابقتين"
        else:
            hashed = hash_pw(pw)
            cur.execute(
                f"UPDATE users SET password_hash={ph()}, reset_token=NULL, reset_expires=NULL WHERE id={ph()}",
                (hashed, u["id"])
            )
            db.commit(); db.close()
            log_action(u["id"], "PASSWORD_RESET", ip=request.remote_addr)
            return render_template("auth_message.html",
                title="تم تغيير كلمة المرور ✓",
                msg="يمكنك الآن تسجيل الدخول بكلمة المرور الجديدة.", type="success")
    db.close()
    return render_template("auth_reset.html", token=token, error=error)


@app.route("/auth/logout")
def auth_logout():
    uid = session.get("user_id")
    if uid: log_action(uid, "LOGOUT", ip=request.remote_addr)
    session.clear()
    return redirect(url_for("auth_login"))


@app.route("/auth/change-password", methods=["POST"])
@login_required
@limiter.limit(lambda: os.getenv("LIMIT_AUTH_CHANGE_PW", "15 per hour"))
def auth_change_password():
    """Allow any authenticated user (superadmin, admin, staff, student) to change their password."""
    uid = session.get("user_id")
    if not uid:
        return jsonify(success=False, message="غير مصرح. يرجى تسجيل الدخول أولاً"), 401

    data = request.get_json() if request.is_json else request.form
    cur_pw  = (data.get("current_password") or "").strip()
    new_pw  = (data.get("new_password") or "").strip()
    new_pw2 = (data.get("confirm_password") or "").strip()

    if not cur_pw:
        return jsonify(success=False, message="يرجى إدخال كلمة المرور الحالية"), 400
    if not new_pw:
        return jsonify(success=False, message="يرجى إدخال كلمة المرور الجديدة"), 400
    if len(new_pw) < 8:
        return jsonify(success=False, message="كلمة المرور الجديدة يجب أن تكون 8 أحرف على الأقل"), 400
    if new_pw != new_pw2:
        return jsonify(success=False, message="كلمتا المرور الجديدتان غير متطابقتين"), 400
    if cur_pw == new_pw:
        return jsonify(success=False, message="كلمة المرور الجديدة يجب أن تختلف عن كلمة المرور الحالية"), 400

    db = get_db(); cur = db.cursor()
    cur.execute(f"SELECT password_hash, email FROM users WHERE id={ph()}", (uid,))
    u = _row_to_dict(cur.fetchone())
    if not u:
        db.close()
        return jsonify(success=False, message="المستخدم غير موجود"), 404

    if not check_pw(cur_pw, u["password_hash"]):
        db.close()
        return jsonify(success=False, message="كلمة المرور الحالية غير صحيحة"), 400

    new_hashed = hash_pw(new_pw)
    cur.execute(f"UPDATE users SET password_hash={ph()} WHERE id={ph()}", (new_hashed, uid))
    db.commit()
    db.close()

    log_action(uid, "CHANGE_PASSWORD", target=u["email"], detail="User changed password", ip=request.remote_addr)
    return jsonify(success=True, message="تم تغيير كلمة المرور بنجاح ✓")


# ══════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════════════════════════════════════════

@app.route("/")
@login_required
def dashboard():
    if session.get("role") == "student":
        sid = session.get("student_id", "")
        if sid:
            db = get_db(); cur = db.cursor()
            cur.execute(f"SELECT student_id FROM students WHERE student_id={ph()}", (sid,))
            exists = cur.fetchone(); db.close()
            if exists:
                return redirect(url_for("student_card", student_id=sid))
        return redirect(url_for("student_self_register"))

    db  = get_db(); cur = db.cursor()
    cur.execute("SELECT COUNT(*) AS c FROM students")
    total = (_row_to_dict(cur.fetchone()) or {}).get("c", 0)
    cur.execute("SELECT COUNT(DISTINCT year) AS c FROM students")
    years = (_row_to_dict(cur.fetchone()) or {}).get("c", 0)
    cur.execute("SELECT COUNT(DISTINCT college) AS c FROM students")
    colleges_n = (_row_to_dict(cur.fetchone()) or {}).get("c", 0)
    # recent audit
    cur.execute("""SELECT a.action, a.target, a.created_at, u.full_name
                   FROM audit_log a LEFT JOIN users u ON a.user_id=u.id
                   ORDER BY a.created_at DESC LIMIT 10""")
    audit = [_row_to_dict(r) for r in cur.fetchall()]
    db.close()
    return render_template("dashboard.html",
        total=total, years=years, colleges_n=colleges_n,
        audit=audit, role=session.get("role"),
        user_name=session.get("user_name"), colleges=COLLEGES,
        all_years=YEAR_RANGE)


# ══════════════════════════════════════════════════════════════════════════
# STUDENT REGISTRATION (staff / admin / superadmin)
# ══════════════════════════════════════════════════════════════════════════

@app.route("/register-form")
@login_required
def register_form():
    role          = session.get("role")
    user_college  = session.get("college")
    user_sid      = session.get("student_id", "")

    # Students can only register themselves — pre-fill their data
    if role == "student":
        if not user_sid:
            return render_template("auth_message.html",
                title="لا يوجد رقم طالب",
                msg="حسابك غير مرتبط برقم طالب. تواصل مع إدارة الكلية.",
                type="error")
        # Check if already registered
        db  = get_db(); cur = db.cursor()
        cur.execute(f"SELECT * FROM students WHERE student_id={ph()}", (user_sid,))
        existing = _row_to_dict(cur.fetchone()); db.close()
        if existing:
            return redirect(url_for("student_card", student_id=user_sid))

    visible_colleges = [user_college] if role == "admin" and user_college else COLLEGES
    # For student role, derive year from student_id prefix
    prefill_year = user_sid[:4] if role == "student" and len(user_sid) >= 4 else ""
    prefill_code = user_sid[4:] if role == "student" and len(user_sid) > 4  else ""

    return render_template("register.html",
                           colleges=visible_colleges,
                           years=YEAR_RANGE,
                           role=role,
                           user_name=session.get("user_name"),
                           locked_college=user_college if role in ("admin","student") else None,
                           locked_student_id=user_sid if role == "student" else None,
                           prefill_year=prefill_year,
                           prefill_code=prefill_code)


@app.route("/register", methods=["POST"])
@login_required
@limiter.limit(lambda: os.getenv("LIMIT_STAFF_REGISTER", "300 per hour"))
def register():
    try:
        full_name   = request.form.get("full_name","").strip()
        year        = to_eng(request.form.get("year","").strip())
        code        = to_eng(request.form.get("code","").strip())
        college     = request.form.get("college","").strip()
        student_email = request.form.get("student_email","").strip().lower()
        rotation    = int(request.form.get("rotation","0"))
        flip_h      = request.form.get("flip_h","") == "1"
        zoom        = float(request.form.get("zoom","1.0"))
        offset_x    = float(request.form.get("offset_x","0.0"))
        offset_y    = float(request.form.get("offset_y","0.0"))
        auto_crop   = request.form.get("auto_crop","1") == "1"
        image_file  = request.files.get("image")

        ok, msg = validate_full_name(full_name)
        if not ok: return jsonify(success=False, message=msg), 400
        ok, msg = validate_student_id(year, code)
        if not ok: return jsonify(success=False, message=msg), 400
        if college not in COLLEGES:
            return jsonify(success=False, message="اختر كلية صحيحة"), 400

        # Admin can only register students in their own college
        if session.get("role") == "admin":
            allowed = session.get("college")
            if allowed and college != allowed:
                return jsonify(success=False,
                    message=f"يمكنك تسجيل طلاب كلية {allowed} فقط"), 403

        student_id = year + code
         # Student can ONLY register themselves
        if session.get("role") == "student":
          my_sid = session.get("student_id", "")
          if student_id != my_sid:
               return jsonify(success=False,
                 message="يمكنك تسجيل بياناتك الشخصية فقط"), 403

        # ── Check duplicate BEFORE processing image ──
        db  = get_db(); cur = db.cursor()
        cur.execute(f"SELECT student_id, full_name, image_path FROM students WHERE student_id={ph()}", (student_id,))
        existing = _row_to_dict(cur.fetchone()); db.close()
        if existing:
            img_url = None
            if existing.get("image_path"):
                p = existing["image_path"]
                img_url = p if p.startswith("http") else f"/static/{p}"
            return jsonify(
                success=False,
                duplicate=True,
                existing_name=existing["full_name"],
                existing_img=img_url,
                existing_id=existing["student_id"],
                message=f"الرقم {student_id} مسجل مسبقاً باسم: {existing['full_name']}"
            ), 409

        if not image_file or image_file.filename == "":
            return jsonify(success=False, message="يرجى رفع صورة شخصية"), 400
        if not image_file.filename.lower().endswith((".jpg",".jpeg")):
            return jsonify(success=False, message="يُسمح فقط بصور JPG"), 400

        raw_bytes = image_file.read()
        if len(raw_bytes) > app.config["MAX_CONTENT_LENGTH"]:
            return jsonify(success=False, message="حجم الصورة يتجاوز 5 MB"), 400

        # Apply edits + auto-crop
        try:
            processed = apply_edits(raw_bytes, rotation=rotation, flip_h=flip_h,
                                    zoom=zoom, offset_x=offset_x, offset_y=offset_y,
                                    auto_crop=auto_crop)
        except Exception as e:
            app.logger.warning(f"Failed to process image: {e}")
            return jsonify(success=False, message="الملف المرفوع ليس صورة صالحة أو أنه تالف."), 400

        # Single-person face check (after processing)
        is_valid, face_msg, _ = validate_single_person(processed)
        if not is_valid:
            return jsonify(success=False, message=face_msg), 400

        # Save
        result = save_image(processed, student_id, year, college, UPLOAD_FOLDER)

        # DB insert
        db  = get_db(); cur = db.cursor()
        try:
            uid = session.get("user_id")
            cur.execute(
                f"INSERT INTO students (student_id,full_name,year,college,email,image_path,registered_by) VALUES ({','.join([ph()]*7)})",
                (student_id, full_name, year, college,
                 student_email or None, result["path"], uid)
            )
            db.commit()
            log_action(uid, "REGISTER_STUDENT", target=student_id,
                       detail=full_name, ip=request.remote_addr)
        except Exception as e:
            db.close()
            # Clean up saved file
            if not result.get("cloudinary") and os.path.exists(
                    os.path.join(STATIC_ROOT, result["path"])):
                os.remove(os.path.join(STATIC_ROOT, result["path"]))
            if "UNIQUE" in str(e) or "unique" in str(e).lower():
                return jsonify(success=False,
                    message=f"⚠️ الرقم {student_id} مسجل مسبقاً", duplicate=True), 409
            raise
        db.close()

        # Send confirmation email if student provided email
        if student_email:
            card_link = url_for("student_card", student_id=student_id, _external=True)
            send_email(student_email, "تم تسجيلك في النظام الجامعي",
                       _email_registered_html(full_name, student_id, card_link))

        return jsonify(success=True,
            message=f"✅ تم تسجيل {full_name} بنجاح! الرقم: {student_id}",
            student_id=student_id,
            card_url=url_for("student_card", student_id=student_id),
            image_url=result["url"]), 201

    except Exception as e:
        app.logger.exception("Register error")
        return jsonify(success=False, message="حدث خطأ داخلي"), 500


# ══════════════════════════════════════════════════════════════════════════
# STUDENT CARD  (public)
# ══════════════════════════════════════════════════════════════════════════

@app.route("/student/<student_id>")
def student_card(student_id):
    student_id = to_eng(student_id.strip())
    db  = get_db(); cur = db.cursor()
    cur.execute(f"SELECT * FROM students WHERE student_id={ph()}", (student_id,))
    row = _row_to_dict(cur.fetchone()); db.close()
    if not row: abort(404)
    return render_template("student_card.html", student=row)


@app.route("/student/<student_id>/update-photo", methods=["POST"])
@limiter.limit(lambda: os.getenv("LIMIT_UPDATE_PHOTO", "30 per hour"))
def update_photo(student_id):
    student_id = to_eng(student_id.strip())
    
    # ── Authentication and Authorization checks (V-006, V-007) ──
    uid = session.get("user_id")
    if not uid:
        return jsonify(success=False, message="غير مصرح بالدخول. يرجى تسجيل الدخول أولاً"), 401
        
    db  = get_db(); cur = db.cursor()
    cur.execute(f"SELECT * FROM students WHERE student_id={ph()}", (student_id,))
    row = _row_to_dict(cur.fetchone())
    if not row:
        db.close()
        return jsonify(success=False, message="غير موجود"), 404

    role = session.get("role")
    my_sid = session.get("student_id")
    my_college = session.get("college")
    
    if role == "student":
        if not my_sid or my_sid != student_id:
            db.close()
            return jsonify(success=False, message="غير مصرح لك بتحديث هذه الصورة"), 403
    elif role == "admin":
        if my_college and row["college"] != my_college:
            db.close()
            return jsonify(success=False, message="غير مصرح لك بتعديل بيانات طالب خارج كليتك"), 403
    elif role not in ("superadmin", "staff"):
        db.close()
        return jsonify(success=False, message="غير مصرح بالدخول"), 403

    image_file = request.files.get("image")
    if not image_file or not image_file.filename.lower().endswith((".jpg",".jpeg")):
        db.close(); return jsonify(success=False, message="يرجى رفع صورة JPG"), 400

    rotation = int(request.form.get("rotation","0"))
    flip_h   = request.form.get("flip_h","") == "1"
    zoom     = float(request.form.get("zoom","1.0"))
    offset_x = float(request.form.get("offset_x","0.0"))
    offset_y = float(request.form.get("offset_y","0.0"))
    auto_crop = request.form.get("auto_crop","1") == "1"

    raw      = image_file.read()
    if len(raw) > app.config["MAX_CONTENT_LENGTH"]:
        db.close(); return jsonify(success=False, message="الصورة أكبر من 5 MB"), 400

    try:
        processed = apply_edits(raw, rotation=rotation, flip_h=flip_h,
                                zoom=zoom, offset_x=offset_x, offset_y=offset_y,
                                auto_crop=auto_crop)
        is_valid, face_msg, _ = validate_single_person(processed)
        if not is_valid:
            db.close()
            return jsonify(success=False, message=face_msg), 400

        # Archive old
        old_rel = row.get("image_path","")
        archived_path = archive_old_image(old_rel, student_id, STATIC_ROOT, UPLOAD_FOLDER)
        if archived_path:
            app.logger.info(f"Archived old image to: {archived_path}")

        # Save new
        result = save_image(processed, student_id, row["year"], row["college"], UPLOAD_FOLDER)

        cur.execute(
            f"UPDATE students SET image_path={ph()}, updated_at={ph()} WHERE student_id={ph()}",
            (result["path"], datetime.utcnow().isoformat(), student_id)
        )
        db.commit()
        uid = session.get("user_id")
        log_action(uid, "UPDATE_PHOTO", target=student_id,
                   detail="photo updated via student card", ip=request.remote_addr)

        db.close()
        return jsonify(success=True, message="تم تحديث الصورة ✓",
                       new_url=result["url"])
    except Exception as e:
        db.close()
        app.logger.error(f"Error updating photo for {student_id}: {e}")
        return jsonify(success=False, message="حدث خطأ أثناء معالجة الصورة. تأكد من أن الملف المرفوع صورة صالحة وغير تالفة."), 500


# ══════════════════════════════════════════════════════════════════════════
# ADMIN – STUDENTS TABLE
# ══════════════════════════════════════════════════════════════════════════

@app.route("/admin/students")
@role_required("superadmin","admin","staff")
def admin_students():
    q        = request.args.get("q","").strip()
    year     = request.args.get("year","").strip()
    college  = request.args.get("college","").strip()
    page     = max(int(request.args.get("page",1)),1)
    per_page = 20
    role     = session.get("role")
    user_college = session.get("college")  # admin restricted to own college

    conditions, params = [], []
    if q:
        if USE_PG:
            conditions.append("(student_id ILIKE %s OR full_name ILIKE %s)")
        else:
            conditions.append("(student_id LIKE ? OR full_name LIKE ?)")
        params += [f"%{q}%", f"%{q}%"]
    if year:    conditions.append(f"year={ph()}");    params.append(year)
    if college: conditions.append(f"college={ph()}"); params.append(college)
    # admin can only see own college
    if role == "admin" and user_college:
        conditions.append(f"college={ph()}"); params.append(user_college)

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    db    = get_db(); cur = db.cursor()
    cur.execute(f"SELECT COUNT(*) AS c FROM students {where}", params)
    total  = (_row_to_dict(cur.fetchone()) or {}).get("c", 0)
    offset = (page-1)*per_page
    if USE_PG:
        cur.execute(f"SELECT * FROM students {where} ORDER BY created_at DESC LIMIT %s OFFSET %s",
                    params+[per_page, offset])
    else:
        cur.execute(f"SELECT * FROM students {where} ORDER BY created_at DESC LIMIT ? OFFSET ?",
                    params+[per_page, offset])
    rows  = [_row_to_dict(r) for r in cur.fetchall()]
    db.close()
    return jsonify(students=rows, total=total, page=page,
                   pages=(total+per_page-1)//per_page)


@app.route("/admin/delete/<int:sid>", methods=["DELETE"])
@role_required("superadmin","admin")
def admin_delete(sid):
    db  = get_db(); cur = db.cursor()
    cur.execute(f"SELECT * FROM students WHERE id={ph()}", (sid,))
    row = _row_to_dict(cur.fetchone())
    if not row: db.close(); return jsonify(success=False, message="غير موجود"), 404
    # Restrict admin to own college
    if session.get("role")=="admin" and row.get("college") != session.get("college"):
        db.close(); return jsonify(success=False, message="ليس لديك صلاحية"), 403
    img_path = os.path.join(STATIC_ROOT, row.get("image_path",""))
    cur.execute(f"DELETE FROM students WHERE id={ph()}", (sid,))
    db.commit(); db.close()
    if os.path.exists(img_path): os.remove(img_path)
    log_action(session.get("user_id"), "DELETE_STUDENT",
               target=row.get("student_id"), detail=row.get("full_name"),
               ip=request.remote_addr)
    return jsonify(success=True, message="تم الحذف")


@app.route("/admin/export")
@role_required("superadmin","admin")
def admin_export():
    db  = get_db(); cur = db.cursor()
    where, params = "", []
    if session.get("role") == "admin" and session.get("college"):
        where  = f"WHERE college={ph()}"
        params = [session.get("college")]
    cur.execute(f"SELECT student_id,full_name,year,college,email,created_at FROM students {where} ORDER BY created_at DESC", params)
    rows = [_row_to_dict(r) for r in cur.fetchall()]; db.close()

    wb = Workbook(); ws = wb.active; ws.title = "الطلاب"
    ws.sheet_view.rightToLeft = True
    hf = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    hb = PatternFill("solid", fgColor="0D1F3C")
    ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
    th = Side(style="thin", color="CCCCCC")
    bd = Border(left=th,right=th,top=th,bottom=th)
    af = PatternFill("solid", fgColor="EBF2FA")
    headers = ["رقم الطالب","الاسم الكامل","السنة","الكلية","الإيمبيل","تاريخ التسجيل"]
    widths  = [18,42,10,36,32,22]
    for ci,(h,w) in enumerate(zip(headers,widths),1):
        cell = ws.cell(1,ci,h); cell.font=hf; cell.fill=hb
        cell.alignment=ca; cell.border=bd
        ws.column_dimensions[cell.column_letter].width=w
    ws.row_dimensions[1].height=28
    for ri,r in enumerate(rows,2):
        for ci,v in enumerate([r.get("student_id"),r.get("full_name"),
                                r.get("year"),r.get("college"),
                                r.get("email",""),r.get("created_at")],1):
            cell=ws.cell(ri,ci,str(v) if v else "")
            cell.alignment=ca; cell.border=bd
            if ri%2==0: cell.fill=af
        ws.row_dimensions[ri].height=22
    ws.freeze_panes="A2"
    ws.auto_filter.ref=f"A1:F{len(rows)+1}"
    tmp=tempfile.NamedTemporaryFile(suffix=".xlsx",delete=False)
    wb.save(tmp.name); tmp.close()
    now=datetime.now().strftime("%Y%m%d_%H%M%S")
    log_action(session.get("user_id"),"EXPORT_EXCEL",ip=request.remote_addr)
    return send_file(tmp.name,as_attachment=True,
                     download_name=f"students_{now}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ══════════════════════════════════════════════════════════════════════════
# USER MANAGEMENT  (superadmin only)
# ══════════════════════════════════════════════════════════════════════════

@app.route("/admin/users")
@role_required("superadmin")
def admin_users():
    db  = get_db(); cur = db.cursor()
    cur.execute("""SELECT id,email,full_name,role,college,student_id,
                          is_active,email_verified,created_at
                   FROM users ORDER BY college NULLS LAST, role, full_name""")
    users = [_row_to_dict(r) for r in cur.fetchall()]
    db.close()

    # Group by college for display
    grouped = {}
    for u in users:
        key = u.get("college") or "إدارة النظام"
        grouped.setdefault(key, []).append(u)

    return jsonify(users=users, grouped={k: v for k, v in sorted(grouped.items())})


@app.route("/admin/users/create", methods=["POST"])
@role_required("superadmin")
def admin_create_user():
    data      = request.get_json() or {}
    email     = (data.get("email") or "").strip().lower()
    role      = data.get("role", "staff")
    college   = data.get("college") or None
    full_name = (data.get("full_name") or "").strip()
    password  = (data.get("password") or "").strip()

    if not full_name or len(full_name.split()) < 2:
        return jsonify(success=False, message="يرجى إدخال الاسم الكامل للمستخدم"), 400
    if not validate_university_email(email):
        return jsonify(success=False, message=f"يجب استخدام بريد الجامعة (@{UNIVERSITY_DOMAIN})"), 400
    if role not in ROLES or role == "student":
        return jsonify(success=False, message="دور غير صالح — الطلاب يسجلون بأنفسهم"), 400
    if not password or len(password) < 8:
        return jsonify(success=False, message="كلمة المرور يجب أن تكون 8 أحرف على الأقل"), 400

    hashed = hash_pw(password)
    db = get_db(); cur = db.cursor()
    try:
        cur.execute(
            f"INSERT INTO users (email,password_hash,full_name,role,college,email_verified,is_active) VALUES ({','.join([ph()]*7)})",
            (email, hashed, full_name, role, college, 1 if not USE_PG else True, 1 if not USE_PG else True)
        )
        db.commit()
    except Exception as e:
        db.close()
        return jsonify(success=False, message="البريد الإلكتروني مسجل مسبقاً"), 409
    db.close()

    try:
        send_email(
            email,
            "تم إنشاء حسابك في نظام التسجيل الجامعي",
            f"""
            <div dir="rtl" style="font-family:Cairo,Arial;max-width:520px;margin:auto">
              <div style="background:#0d1f3c;padding:28px;border-radius:14px 14px 0 0;text-align:center">
                <h2 style="color:#e8b84b;margin:0">مرحباً بك في نظام التسجيل</h2>
              </div>
              <div style="background:#f0f4f9;padding:28px;border-radius:0 0 14px 14px">
                <p>أهلاً <strong>{full_name}</strong>،</p>
                <p>تم إنشاء حساب لك بدور: <strong>{ROLES.get(role, role)}</strong>.</p>
                <p>يمكنك الآن تسجيل الدخول باستخدام بريدك الجامعي وكلمة المرور المحددة لك من قبل الإدارة.</p>
                <p style="color:#6b7a99;font-size:.85rem">يمكنك تغيير كلمة مرورك في أي وقت بسهولة من داخل حسابك بعد تسجيل الدخول.</p>
              </div>
            </div>
            """
        )
    except Exception:
        pass

    log_action(session.get("user_id"), "CREATE_USER", target=email, detail=f"Role: {role}", ip=request.remote_addr)
    return jsonify(success=True, message=f"تم إنشاء حساب {full_name} بنجاح ويمكنه تسجيل الدخول بكلمة المرور المحددة")


@app.route("/admin/users/<int:uid>/reset-password", methods=["POST"])
@role_required("superadmin")
def admin_reset_user_password(uid):
    """Allow superadmin to set/reset any user's password directly."""
    data = request.get_json() if request.is_json else request.form
    new_pw = (data.get("new_password") or "").strip()
    if not new_pw or len(new_pw) < 8:
        return jsonify(success=False, message="كلمة المرور الجديدة يجب أن تكون 8 أحرف على الأقل"), 400

    db = get_db(); cur = db.cursor()
    cur.execute(f"SELECT email, full_name FROM users WHERE id={ph()}", (uid,))
    u = _row_to_dict(cur.fetchone())
    if not u:
        db.close()
        return jsonify(success=False, message="المستخدم غير موجود"), 404

    new_hashed = hash_pw(new_pw)
    cur.execute(f"UPDATE users SET password_hash={ph()} WHERE id={ph()}", (new_hashed, uid))
    db.commit()
    db.close()

    log_action(session.get("user_id"), "ADMIN_RESET_PASSWORD", target=u["email"], detail=f"Reset password for {u['full_name']}", ip=request.remote_addr)
    return jsonify(success=True, message=f"تم تعيين كلمة المرور الجديدة للمستخدم {u['full_name']} بنجاح ✓")


@app.route("/admin/users/<int:uid>/toggle", methods=["POST"])
@role_required("superadmin")
def admin_toggle_user(uid):
    db = get_db(); cur = db.cursor()
    cur.execute(f"SELECT is_active,email FROM users WHERE id={ph()}", (uid,))
    u = _row_to_dict(cur.fetchone())
    if not u: db.close(); return jsonify(success=False), 404
    new_val = (not u["is_active"]) if USE_PG else (0 if u["is_active"] else 1)
    cur.execute(f"UPDATE users SET is_active={ph()} WHERE id={ph()}", (new_val, uid))
    db.commit(); db.close()
    log_action(session.get("user_id"), "TOGGLE_USER", target=u["email"], ip=request.remote_addr)
    return jsonify(success=True, active=bool(new_val))


# ══════════════════════════════════════════════════════════════════════════
# AUDIT LOG
# ══════════════════════════════════════════════════════════════════════════

@app.route("/admin/audit")
@role_required("superadmin")
def admin_audit():
    page     = max(int(request.args.get("page",1)),1)
    per_page = 50
    offset   = (page-1)*per_page
    db = get_db(); cur = db.cursor()
    cur.execute("SELECT COUNT(*) AS c FROM audit_log")
    total = (_row_to_dict(cur.fetchone()) or {}).get("c",0)
    if USE_PG:
        cur.execute("""SELECT a.*,u.full_name,u.email FROM audit_log a
                       LEFT JOIN users u ON a.user_id=u.id
                       ORDER BY a.created_at DESC LIMIT %s OFFSET %s""", (per_page,offset))
    else:
        cur.execute("""SELECT a.*,u.full_name,u.email FROM audit_log a
                       LEFT JOIN users u ON a.user_id=u.id
                       ORDER BY a.created_at DESC LIMIT ? OFFSET ?""", (per_page,offset))
    rows = [_row_to_dict(r) for r in cur.fetchall()]; db.close()
    return jsonify(logs=rows, total=total, page=page,
                   pages=(total+per_page-1)//per_page)


# ══════════════════════════════════════════════════════════════════════════
# JWT API  (for external integrations)
# ══════════════════════════════════════════════════════════════════════════

@app.route("/api/login", methods=["POST"])
@csrf.exempt
@limiter.limit(lambda: os.getenv("LIMIT_API_LOGIN", "60 per hour"))
def api_login():
    data  = request.get_json() or {}
    email = data.get("email","").strip().lower()
    pw    = data.get("password","")
    db    = get_db(); cur = db.cursor()
    cur.execute(f"SELECT * FROM users WHERE email={ph()}", (email,))
    u = _row_to_dict(cur.fetchone()); db.close()
    if not u or not u.get("is_active") or not u.get("email_verified"):
        return jsonify(msg="بيانات الدخول غير صحيحة"), 401
    if not check_pw(pw, u["password_hash"]):
        return jsonify(msg="بيانات الدخول غير صحيحة"), 401
    token = create_access_token(identity={"id":u["id"],"role":u["role"],"email":u["email"]})
    log_action(u["id"],"API_LOGIN",ip=request.remote_addr)
    return jsonify(access_token=token, role=u["role"])


@app.route("/api/students")
@jwt_required()
def api_students():
    page     = max(int(request.args.get("page",1)),1)
    per_page = 20; offset=(page-1)*per_page
    db = get_db(); cur = db.cursor()
    if USE_PG:
        cur.execute("SELECT student_id,full_name,year,college,image_path,created_at FROM students ORDER BY created_at DESC LIMIT %s OFFSET %s", (per_page,offset))
    else:
        cur.execute("SELECT student_id,full_name,year,college,image_path,created_at FROM students ORDER BY created_at DESC LIMIT ? OFFSET ?", (per_page,offset))
    rows=[_row_to_dict(r) for r in cur.fetchall()]
    cur.execute("SELECT COUNT(*) AS c FROM students")
    total=(_row_to_dict(cur.fetchone()) or {}).get("c",0)
    db.close()
    return jsonify(students=rows, total=total, page=page)


@app.route("/api/students/<student_id>")
@jwt_required()
def api_student(student_id):
    db=get_db(); cur=db.cursor()
    cur.execute(f"SELECT * FROM students WHERE student_id={ph()}",(to_eng(student_id),))
    row=_row_to_dict(cur.fetchone()); db.close()
    if not row: return jsonify(msg="غير موجود"),404
    row.pop("id",None)
    return jsonify(row)


# ══════════════════════════════════════════════════════════════════════════
# MAIN ADMIN PAGE  (renders the SPA shell)
# ══════════════════════════════════════════════════════════════════════════

@app.route("/admin")
@role_required("superadmin","admin","staff")
def admin_panel():
    db=get_db(); cur=db.cursor()
    yrs=[_row_to_dict(r)["year"]    for r in cur.execute("SELECT DISTINCT year FROM students ORDER BY year DESC").fetchall()] if not USE_PG else \
        [r["year"] for r in (cur.execute("SELECT DISTINCT year FROM students ORDER BY year DESC") or [cur.fetchall()])[0]] if False else []
    # simpler approach
    cur.execute("SELECT DISTINCT year FROM students ORDER BY year DESC")
    yrs=[_row_to_dict(r)["year"] for r in cur.fetchall()]
    db.close()
    return render_template("admin_panel.html",
        years=yrs, all_years=YEAR_RANGE,
        colleges=COLLEGES, roles=ROLES,
        role=session.get("role"),
        user_name=session.get("user_name"),
        user_college=session.get("college"),
        domain=UNIVERSITY_DOMAIN)


# ══════════════════════════════════════════════════════════════════════════
# BULK IMPORT  –  superadmin uploads Excel, system creates student accounts
# ══════════════════════════════════════════════════════════════════════════

@app.route("/admin/bulk-import", methods=["GET", "POST"])
@role_required("superadmin", "admin")
def bulk_import():
    if request.method == "GET":
        return render_template("bulk_import.html",
            role=session.get("role"),
            user_name=session.get("user_name"),
            domain=UNIVERSITY_DOMAIN)

    # POST — process uploaded Excel/CSV
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify(success=False, message="يرجى رفع ملف Excel أو CSV"), 400

    ext = f.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("xlsx", "xls", "csv"):
        return jsonify(success=False, message="يُقبل xlsx أو csv فقط"), 400

    import io as _io
    raw = f.read()

    rows = []
    try:
        if ext == "csv":
            import csv
            text   = raw.decode("utf-8-sig")
            reader = csv.DictReader(_io.StringIO(text))
            rows   = list(reader)
        else:
            from openpyxl import load_workbook
            wb     = load_workbook(_io.BytesIO(raw), read_only=True, data_only=True)
            ws     = wb.active
            headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for xl_row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, [str(v or "").strip() for v in xl_row])))
    except Exception as e:
        return jsonify(success=False, message=f"خطأ في قراءة الملف: {e}"), 400

    # Expected columns (flexible mapping)
    COL_MAP = {
        "student_id": ["student_id","رقم_الطالب","رقم الطالب","id","الرقم"],
        "full_name":  ["full_name","الاسم_الكامل","الاسم الكامل","name","الاسم"],
        "year":       ["year","السنة","العام","سنة"],
        "college":    ["college","الكلية","كلية"],
        "email":      ["email","الايميل","البريد","البريد_الإلكتروني","ايميل"],
    }

    def find_col(row_dict, aliases):
        for a in aliases:
            if a in row_dict: return row_dict[a]
        return ""

    results = {"created": 0, "skipped": 0, "errors": [], "preview": []}

    for i, row in enumerate(rows[:500], start=2):  # max 500 rows
        sid       = to_eng(find_col(row, COL_MAP["student_id"]).strip())
        full_name = find_col(row, COL_MAP["full_name"]).strip()
        year      = to_eng(find_col(row, COL_MAP["year"]).strip())
        college   = find_col(row, COL_MAP["college"]).strip()
        email     = find_col(row, COL_MAP["email"]).strip().lower()

        if not sid or not full_name:
            results["errors"].append(f"سطر {i}: رقم الطالب أو الاسم مفقود")
            results["skipped"] += 1
            continue

        # Validate college
        if college and college not in COLLEGES:
            # Try partial match
            matched = next((c for c in COLLEGES if college in c or c in college), None)
            if matched:
                college = matched
            else:
                college = COLLEGES[0]  # fallback

        # admin restricted to own college
        if session.get("role") == "admin" and session.get("college"):
            college = session.get("college")

        if not year or not re.fullmatch(r"\d{4}", year):
            year = sid[:4] if len(sid) >= 4 else str(CURRENT_YEAR)

        # Check if already exists
        db  = get_db(); cur = db.cursor()
        cur.execute(f"SELECT id FROM students WHERE student_id={ph()}", (sid,))
        if cur.fetchone():
            db.close()
            results["skipped"] += 1
            results["preview"].append({"sid": sid, "name": full_name, "status": "موجود مسبقاً"})
            continue

        # Generate temp password
        temp_pw     = secrets.token_urlsafe(8)
        placeholder_img = os.path.join(UPLOAD_FOLDER, "placeholder.jpg")

        # Create a placeholder image if needed
        if not os.path.exists(placeholder_img):
            try:
                from PIL import Image, ImageDraw
                img_ph = Image.new("RGB", (400, 500), color=(26, 58, 107))
                draw   = ImageDraw.Draw(img_ph)
                draw.rectangle([160, 100, 240, 180], fill=(232, 184, 75))
                img_ph.save(placeholder_img, "JPEG")
            except Exception:
                pass

        # Save placeholder in correct folder
        from image_processor import _college_folder
        col_folder = _college_folder(college)
        img_dir    = os.path.join(UPLOAD_FOLDER, year, col_folder)
        os.makedirs(img_dir, exist_ok=True)
        img_name   = f"{sid}_pending.jpg"
        img_path   = os.path.join(img_dir, img_name)
        rel_path   = f"uploads/{year}/{col_folder}/{img_name}"

        if os.path.exists(placeholder_img):
            import shutil
            shutil.copy2(placeholder_img, img_path)
        else:
            with open(img_path, "wb") as fh:
                fh.write(b"")

        # Insert student record
        try:
            uid = session.get("user_id")
            cur.execute(
                f"INSERT INTO students (student_id,full_name,year,college,email,image_path,registered_by) VALUES ({','.join([ph()]*7)})",
                (sid, full_name, year, college, email or None, rel_path, uid)
            )
            db.commit()

            # ── Create student user account ──
            # Email = student_id@domain  (e.g. 2024001001@university.edu.eg)
            student_login_email = f"{sid}@{UNIVERSITY_DOMAIN}"
            temp_pw             = secrets.token_urlsafe(10)
            hashed_pw           = hash_pw(temp_pw)

            cur.execute(f"SELECT id FROM users WHERE email={ph()}", (student_login_email,))
            if not cur.fetchone():
                cur.execute(
                    f"INSERT INTO users (email,password_hash,full_name,role,college,student_id,is_active,email_verified) VALUES ({','.join([ph()]*8)})",
                    (student_login_email, hashed_pw, full_name, "student",
                     college, sid, 1, 1)   # pre-verified, active
                )
                db.commit()

            results["created"] += 1
            results["preview"].append({
                "sid":      sid,
                "name":     full_name,
                "status":   "تم الإنشاء",
                "email":    student_login_email,
                "temp_pw":  temp_pw,
            })
            log_action(uid, "BULK_IMPORT_STUDENT", target=sid,
                       detail=full_name, ip=request.remote_addr)
        except Exception as e:
            results["errors"].append(f"سطر {i} ({sid}): {e}")
            results["skipped"] += 1
            results["preview"].append({"sid": sid, "name": full_name, "status": "خطأ"})
        finally:
            db.close()

        # Send welcome email with login credentials
        if email:
            card_link   = url_for("student_card", student_id=sid, _external=True)
            login_email = f"{sid}@{UNIVERSITY_DOMAIN}"
            _send_student_welcome(email, full_name, sid, login_email, temp_pw, card_link)

    return jsonify(success=True, results=results)


@app.route("/admin/bulk-import/template")
@role_required("superadmin", "admin")
def bulk_import_template():
    """Download a sample Excel template for bulk import."""
    wb = Workbook()
    ws = wb.active
    ws.title = "بيانات الطلاب"
    ws.sheet_view.rightToLeft = True

    from openpyxl.styles import Font, PatternFill, Alignment
    hf = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    hb = PatternFill("solid", fgColor="0D1F3C")
    ca = Alignment(horizontal="center", vertical="center")

    headers = ["student_id", "full_name", "year", "college", "email"]
    ar_headers = ["رقم الطالب", "الاسم الكامل", "السنة", "الكلية", "الإيميل"]
    widths = [18, 40, 8, 36, 34]

    for ci, (h, ah, w) in enumerate(zip(headers, ar_headers, widths), 1):
        cell = ws.cell(1, ci, f"{h} / {ah}")
        cell.font = hf; cell.fill = hb; cell.alignment = ca
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 26

    # Sample rows
    samples = [
        ("2024001001", "محمد أحمد علي حسن",      "2024", "كلية الحاسبات والمعلومات", "student1@university.edu.eg"),
        ("2024001002", "فاطمة عبدالله إبراهيم سيد","2024", "كلية الطب البشري",         ""),
        ("2023005010", "أحمد محمود خالد عمر",      "2023", "كلية الهندسة",             "student3@university.edu.eg"),
    ]
    from openpyxl.styles import PatternFill as PF
    alt = PF("solid", fgColor="EBF2FA")
    for ri, row in enumerate(samples, 2):
        for ci, v in enumerate(row, 1):
            cell = ws.cell(ri, ci, v)
            cell.alignment = ca
            if ri % 2 == 0: cell.fill = alt
        ws.row_dimensions[ri].height = 20

    ws.freeze_panes = "A2"
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name); tmp.close()
    return send_file(tmp.name, as_attachment=True,
                     download_name="students_import_template.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



def _send_student_welcome(to: str, name: str, student_id: str,
                          login_email: str, temp_pw: str, card_link: str):
    """Send welcome email with login credentials to newly imported student."""
    html = f"""
<div dir="rtl" style="font-family:Cairo,Arial;max-width:540px;margin:auto">
  <div style="background:linear-gradient(135deg,#0d1f3c,#1a3a6b);padding:30px;border-radius:14px 14px 0 0;text-align:center">
    <h1 style="color:#e8b84b;margin:0;font-size:1.4rem">&#127891; مرحباً بك في النظام الجامعي</h1>
  </div>
  <div style="background:#f0f4f9;padding:28px;border-radius:0 0 14px 14px">
    <p style="font-size:1rem">أهلاً <strong>{name}</strong>،</p>
    <p style="margin-top:8px;color:#444">تم تسجيلك في نظام القيد الجامعي.</p>

    <div style="background:#0d1f3c;border-radius:10px;padding:16px;margin:16px 0">
      <p style="color:rgba(255,255,255,.5);font-size:.8rem;margin-bottom:8px;text-align:center">رقمك الجامعي</p>
      <p style="color:#e8b84b;font-family:monospace;font-size:1.5rem;font-weight:700;letter-spacing:3px;text-align:center">{student_id}</p>
    </div>

    <div style="background:#fff;border:1px solid #dce3ef;border-radius:10px;padding:16px;margin:14px 0">
      <p style="font-weight:700;color:#1a2744;margin-bottom:10px">بيانات تسجيل الدخول:</p>
      <p style="font-size:.88rem;color:#444;margin-bottom:5px">البريد: <strong style="direction:ltr;display:inline-block">{login_email}</strong></p>
      <p style="font-size:.88rem;color:#444">كلمة المرور المؤقتة: <strong style="font-family:monospace;letter-spacing:1px">{temp_pw}</strong></p>
      <p style="font-size:.75rem;color:#c53030;margin-top:8px">* يُنصح بتغيير كلمة المرور بعد أول دخول</p>
    </div>

    <a href="{card_link}" style="display:inline-block;background:#1a3a6b;color:#e8b84b;
      padding:12px 28px;border-radius:8px;text-decoration:none;font-weight:700;font-size:1rem;margin:10px 0">
      فتح بطاقة الهوية ورفع الصورة
    </a>
  </div>
</div>"""
    send_email(to, "مرحباً – تم تسجيلك في النظام الجامعي", html)


# ══════════════════════════════════════════════════════════════════════════
# STUDENT SELF-REGISTRATION  (student role only — simple dedicated page)
# ══════════════════════════════════════════════════════════════════════════

@app.route("/student/register", methods=["GET"])
@login_required
def student_self_register():
    """Dedicated registration page for students — only their own data."""
    if session.get("role") != "student":
        return redirect(url_for("register_form"))

    sid = session.get("student_id", "")
    # Already registered → go to card
    if sid:
        db  = get_db(); cur = db.cursor()
        cur.execute(f"SELECT student_id FROM students WHERE student_id={ph()}", (sid,))
        if cur.fetchone():
            db.close()
            return redirect(url_for("student_card", student_id=sid))
        db.close()

    return render_template("student_self_register.html",
                           colleges=COLLEGES,
                           years=YEAR_RANGE,
                           user_name=session.get("user_name"),
                           prefill_sid=sid)


@app.route("/student/register", methods=["POST"])
@login_required
@limiter.limit(lambda: os.getenv("LIMIT_STUDENT_REGISTER", "30 per hour"))
def student_self_register_post():
    """Handle student self-registration form submission."""
    if session.get("role") != "student":
        return jsonify(success=False, message="غير مسموح"), 403
    try:
        full_name  = request.form.get("full_name","").strip()
        year       = to_eng(request.form.get("year","").strip())
        code       = to_eng(request.form.get("code","").strip())
        college    = request.form.get("college","").strip()
        rotation   = int(request.form.get("rotation","0"))
        flip_h     = request.form.get("flip_h","") == "1"
        zoom       = float(request.form.get("zoom","1.0"))
        offset_x   = float(request.form.get("offset_x","0.0"))
        offset_y   = float(request.form.get("offset_y","0.0"))
        auto_crop  = request.form.get("auto_crop","1") == "1"
        image_file = request.files.get("image")

        ok, msg = validate_full_name(full_name)
        if not ok: return jsonify(success=False, message=msg), 400
        ok, msg = validate_student_id(year, code)
        if not ok: return jsonify(success=False, message=msg), 400
        if college not in COLLEGES:
            return jsonify(success=False, message="اختر كلية صحيحة"), 400

        student_id = year + code

        # Enforce: student can only register their own ID
        my_sid = session.get("student_id", "")
        if my_sid and student_id != my_sid:
            return jsonify(success=False,
                message=f"يمكنك تسجيل رقمك الجامعي فقط ({my_sid})"), 403

        # Check duplicate
        db  = get_db(); cur = db.cursor()
        cur.execute(f"SELECT student_id, full_name, image_path FROM students WHERE student_id={ph()}", (student_id,))
        existing = _row_to_dict(cur.fetchone()); db.close()
        if existing:
            img_url = None
            if existing.get("image_path"):
                p = existing["image_path"]
                img_url = p if p.startswith("http") else f"/static/{p}"
            return jsonify(
                success=False, duplicate=True,
                existing_name=existing["full_name"],
                existing_img=img_url,
                existing_id=existing["student_id"],
                message=f"الرقم {student_id} مسجل مسبقاً باسم: {existing['full_name']}"
            ), 409

        if not image_file or not image_file.filename.lower().endswith((".jpg",".jpeg")):
            return jsonify(success=False, message="يرجى رفع صورة JPG"), 400

        raw = image_file.read()
        if len(raw) > app.config["MAX_CONTENT_LENGTH"]:
            return jsonify(success=False, message="حجم الصورة يتجاوز 5 MB"), 400

        try:
            processed = apply_edits(raw, rotation=rotation, flip_h=flip_h,
                                    zoom=zoom, offset_x=offset_x, offset_y=offset_y,
                                    auto_crop=auto_crop)
        except Exception as e:
            app.logger.warning(f"Failed to process student self-registration image: {e}")
            return jsonify(success=False, message="الملف المرفوع ليس صورة صالحة أو أنه تالف."), 400
        is_valid, face_msg, _ = validate_single_person(processed)
        if not is_valid:
            return jsonify(success=False, message=face_msg), 400

        result = save_image(processed, student_id, year, college, UPLOAD_FOLDER)

        db  = get_db(); cur = db.cursor()
        try:
            uid = session.get("user_id")
            cur.execute(
                f"INSERT INTO students (student_id,full_name,year,college,email,image_path,registered_by) VALUES ({','.join([ph()]*7)})",
                (student_id, full_name, year, college,
                 session.get("email"), result["path"], uid)
            )
            # Link student_id to user account
            cur.execute(f"UPDATE users SET student_id={ph()} WHERE id={ph()}",
                        (student_id, uid))
            db.commit()
            log_action(uid, "STUDENT_SELF_REGISTER", target=student_id,
                       detail=full_name, ip=request.remote_addr)
        except Exception as e:
            db.close()
            if "UNIQUE" in str(e) or "unique" in str(e).lower():
                return jsonify(success=False,
                    message=f"الرقم {student_id} مسجل مسبقاً", duplicate=True), 409
            raise
        db.close()

        # Update session
        session["student_id"] = student_id

        return jsonify(success=True,
            message=f"تم تسجيل بياناتك بنجاح!",
            card_url=url_for("student_card", student_id=student_id),
            image_url=result["url"]), 201

    except Exception:
        app.logger.exception("Student self-register error")
        return jsonify(success=False, message="حدث خطأ داخلي"), 500


# ══════════════════════════════════════════════════════════════════════════
# ONEDRIVE OAUTH GATEWAY
# ══════════════════════════════════════════════════════════════════════════

@app.route("/admin/onedrive/auth")
@login_required
def onedrive_auth():
    if session.get("role") != "superadmin":
        abort(403)
    
    client_id = os.getenv("ONEDRIVE_CLIENT_ID")
    if not client_id:
        return "Error: ONEDRIVE_CLIENT_ID is not configured in .env", 400
        
    redirect_uri = f"{request.scheme}://{request.host}/admin/onedrive/callback"
    tenant_id = os.getenv("ONEDRIVE_TENANT_ID", "common")
    
    auth_url = (
        f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/authorize"
        f"?client_id={client_id}"
        f"&response_type=code"
        f"&redirect_uri={redirect_uri}"
        f"&response_mode=query"
        f"&scope=Files.ReadWrite.All%20offline_access"
    )
    return redirect(auth_url)


@app.route("/admin/onedrive/callback")
@login_required
def onedrive_callback():
    if session.get("role") != "superadmin":
        abort(403)
        
    code = request.args.get("code")
    if not code:
        err = request.args.get("error")
        err_desc = request.args.get("error_description")
        if err or err_desc:
            return f"Microsoft OAuth Error: {err}<br>Description: {err_desc}", 400
        return f"Error: Authorization code is missing from query string. Received query parameters: {dict(request.args)}", 400
        
    client_id = os.getenv("ONEDRIVE_CLIENT_ID")
    client_secret = os.getenv("ONEDRIVE_CLIENT_SECRET")
    tenant_id = os.getenv("ONEDRIVE_TENANT_ID", "common")
    
    if not client_id or not client_secret:
        return "Error: Client ID or Client Secret is not configured in .env", 400
        
    redirect_uri = f"{request.scheme}://{request.host}/admin/onedrive/callback"
    token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    
    data = {
        "client_id": client_id,
        "client_secret": client_secret,
        "code": code,
        "redirect_uri": redirect_uri,
        "grant_type": "authorization_code",
        "scope": "Files.ReadWrite.All offline_access"
    }
    
    try:
        import requests
        res = requests.post(token_url, data=data, timeout=15)
        if res.status_code == 200:
            tokens = res.json()
            refresh_token = tokens.get("refresh_token")
            
            html = f"""
            <div dir="rtl" style="font-family:Cairo,Arial,sans-serif;max-width:600px;margin:50px auto;padding:30px;border:1px solid #dce3ef;border-radius:14px;background:#f0f4f9;box-shadow:0 8px 30px rgba(0,0,0,0.05)">
              <h1 style="color:#2b6cb0;margin-top:0">🎉 تم الاتصال بـ Microsoft OneDrive بنجاح!</h1>
              <p style="color:#4a5568;line-height:1.6">تم الحصول على رمز التحديث (Refresh Token) بنجاح. يرجى نسخه ووضعه في ملف <strong>.env</strong> الخاص بالتطبيق:</p>
              
              <div style="background:#2d3748;color:#fff;padding:16px;border-radius:8px;font-family:monospace;font-size:0.9rem;word-break:break-all;margin:20px 0;user-select:all" title="انقر لتحديد الكل">
                ONEDRIVE_REFRESH_TOKEN={refresh_token}
              </div>
              
              <p style="color:#e53e3e;font-size:0.85rem;font-weight:bold">* تنبيه: هذا الرمز سري للغاية ويسمح بالوصول لملفاتك، لا تشاركه مع أي شخص.</p>
              <p style="color:#718096;font-size:0.8rem">بعد تعديل ملف .env، أعد تشغيل السيرفر لتفعيل مزامنة الصور تلقائياً.</p>
              <a href="/" style="display:inline-block;margin-top:20px;padding:10px 20px;background:#3182ce;color:#fff;border-radius:8px;text-decoration:none;font-weight:bold">الذهاب للوحة التحكم</a>
            </div>
            """
            return html
        else:
            return f"Error exchanging code: {res.status_code} - {res.text}", 400
    except Exception as e:
        return f"Connection error: {e}", 500


@app.errorhandler(403)
def forbidden(e):
    return render_template("auth_message.html",
        title="ليس لديك صلاحية",
        msg="هذه الصفحة تتطلب صلاحيات أعلى.", type="error"), 403

@app.errorhandler(404)
def not_found(e):
    return render_template("auth_message.html",
        title="الصفحة غير موجودة",
        msg="تأكد من الرابط وحاول مجدداً.", type="error"), 404


# ── run ────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    init_db()
    app.run(debug=False, host="0.0.0.0", port=5000)